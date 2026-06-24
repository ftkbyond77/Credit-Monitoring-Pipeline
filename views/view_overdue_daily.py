# views/view_overdue_daily.py
#
# Jelly Section: Overdue Daily
# แก้ไข: HTML report
#   - ใช้ tempfile.NamedTemporaryFile + webbrowser.open() เหมือน example
#   - ตารางไม่มี CSS — ปล่อย browser default เพื่อให้ paste ใน Outlook ได้ line table
#   - หน้าเว็บ minimal styling เฉพาะ layout ไม่แตะ table

import base64
import tempfile
import urllib.parse
import webbrowser
from datetime import date
from io import BytesIO

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd

from components import dash_title_bar, section_header


# ---------------------------------------------------------------------------
# Default config
# ---------------------------------------------------------------------------
DEFAULT_RECIPIENT  = "nichapas@scg.com"
DEFAULT_SALUTATION = "Dear P'Nong ka"
DEFAULT_CLOSING    = "ขอบคุณค่ะ\nเจล"
DEFAULT_SIGNATURE  = (
    "Best Regards,\n"
    "*-**-**-**-**-**-**-**-**-**-**-*\n"
    "Kanokporn  Yaemklad (JEL)\n"
    "Thai MMA Company Limited (TMMA)\n"
    "MOBILE 0941327798\n"
    "E-MAIL admin_mk3@scg.com"
)

TARGET_COMPANY = "1190"


# ---------------------------------------------------------------------------
# Dynamic subject
# ---------------------------------------------------------------------------
def _build_subject(target_date: date) -> str:
    month_abbr = target_date.strftime("%b").upper()
    year       = target_date.strftime("%Y")
    return f"RE: Over Due {month_abbr} {year}"


# ---------------------------------------------------------------------------
# Parse CollectionDate string "YYYYMMDD"
# ---------------------------------------------------------------------------
def _parse_collection_date(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "CollectionDate" in df.columns:
        df["CollectionDate"] = (
            df["CollectionDate"]
            .astype(str).str.strip()
            .replace({"nan": "", "NaT": ""})
        )
        df["_CollDate_parsed"] = pd.to_datetime(
            df["CollectionDate"], format="%Y%m%d", errors="coerce"
        ).dt.date
    return df


# ---------------------------------------------------------------------------
# Auto-Filter
# ---------------------------------------------------------------------------
def _apply_auto_filter(df: pd.DataFrame, target_date: date) -> pd.DataFrame:
    df = _parse_collection_date(df)

    if "CompanyCode" in df.columns:
        df["CompanyCode"] = df["CompanyCode"].astype(str).str.strip()
    if "OverdueAmount" in df.columns:
        df["OverdueAmount"] = pd.to_numeric(df["OverdueAmount"], errors="coerce").fillna(0.0)

    mask = (
        (df.get("CompanyCode",        pd.Series(dtype=str))    == TARGET_COMPANY)
        & (df.get("_CollDate_parsed", pd.Series(dtype=object)) == target_date)
        & (df.get("OverdueAmount",    pd.Series(dtype=float))  > 0)
    )
    return df[mask].drop(columns=["_CollDate_parsed"], errors="ignore").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Export .xlsx
# ---------------------------------------------------------------------------
def _build_excel(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill(start_color="1A7A4A", end_color="1A7A4A", fill_type="solid")
    HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
    HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    THIN_SIDE     = Side(style="thin", color="C8CDD4")
    CELL_BORDER   = Border(
        left=THIN_SIDE, right=THIN_SIDE,
        top=THIN_SIDE,  bottom=THIN_SIDE,
    )

    def _style_sheet(ws):
        # --- style header row ---
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border    = CELL_BORDER

        # --- border + alignment ทุก data row ---
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border    = CELL_BORDER
                cell.alignment = Alignment(vertical="center")

        # --- auto column width ---
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # --- freeze header row ---
        ws.freeze_panes = "A2"

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Summary sheet
        df.to_excel(writer, sheet_name="Summary", index=False)
        _style_sheet(writer.sheets["Summary"])

        # Sheet ต่อ CustomerName
        if "CustomerName" in df.columns:
            for customer, group in df.groupby("CustomerName"):
                sheet_name = str(customer)[:31]
                group.reset_index(drop=True).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
                _style_sheet(writer.sheets[sheet_name])

    return output.getvalue()

# ---------------------------------------------------------------------------
# Build report HTML
# KEY: ตารางไม่มี CSS เลย — ใช้ border="1" attribute เท่านั้น
#      เพื่อให้ paste ใน Outlook ได้ line table โครงสร้างเดิม
# ---------------------------------------------------------------------------
def _build_report_html(
    filtered_df: pd.DataFrame,
    target_date: date,
    email_body_text: str,
) -> str:
    report_date_str = target_date.strftime("%d %B %Y")

    if "CustomerName" in filtered_df.columns:
        customers = sorted(filtered_df["CustomerName"].dropna().unique())
    else:
        customers = ["All Records"]

    # สร้าง customer block HTML
    customer_blocks_html = ""
    for i, customer in enumerate(customers):
        if "CustomerName" in filtered_df.columns:
            group = filtered_df[filtered_df["CustomerName"] == customer].reset_index(drop=True)
        else:
            group = filtered_df.reset_index(drop=True)

        total      = float(group["OverdueAmount"].sum()) if "OverdueAmount" in group.columns else 0.0
        section_id = f"report_{i}"
        table_html = group.to_html(index=False, border=1, classes="")

        customer_blocks_html += f"""
        <div class="customer-block" id="block_{i}">
            <div class="block-header">
                <div class="block-header-left">
                    <input
                        type="checkbox"
                        class="row-check"
                        data-index="{i}"
                        id="chk_{i}"
                        title="Select for Copy Multi"
                    />
                    <label for="chk_{i}" class="customer-title">{customer}</label>
                    <span class="block-meta">
                        {len(group)} records &nbsp;|&nbsp; OverdueAmount: {total:,.2f} THB
                    </span>
                </div>
                <button
                    class="btn-copy"
                    onclick="copySingle('{section_id}', this)"
                >
                    Copy Table
                </button>
            </div>
            <div class="table-scroll" id="{section_id}">
                {table_html}
            </div>
        </div>
        """

    sig_escaped = email_body_text.replace("<", "&lt;").replace(">", "&gt;")

    return f"""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Overdue Daily Report — {report_date_str}</title>
<style>

  /* ============================================================
     Reset + Base
  ============================================================ */
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    font-size: 13px;
    background: #eef1f6;
    color: #1e2d45;
    padding: 28px 32px;
  }}

  /* ============================================================
     Page Header
  ============================================================ */
  .page-header {{
    background: linear-gradient(120deg, #2c4a7c 0%, #3d6499 100%);
    color: white;
    border-radius: 10px;
    padding: 22px 28px;
    margin-bottom: 20px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }}
  .page-header h1 {{
    font-size: 1.15rem;
    font-weight: 700;
    letter-spacing: 0.01em;
    margin-bottom: 4px;
  }}
  .page-header .meta {{
    font-size: 0.76rem;
    opacity: 0.78;
  }}
  .badge {{
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.35);
    border-radius: 20px;
    padding: 5px 16px;
    white-space: nowrap;
  }}

  /* ============================================================
     Email Preview Box
  ============================================================ */
  .email-box {{
    background: white;
    border-left: 3px solid #3d6499;
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 20px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.07);
  }}
  .email-box .label {{
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #3d6499;
    margin-bottom: 10px;
  }}
  .email-box pre {{
    font-family: 'Segoe UI', sans-serif;
    font-size: 0.85rem;
    line-height: 1.75;
    white-space: pre-wrap;
    color: #1e2d45;
  }}
  .sig {{ color: #b83232; font-weight: 500; }}

  /* ============================================================
     Action Bar (Copy All + Copy Multi)
  ============================================================ */
  .action-bar {{
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 16px;
    padding: 12px 16px;
    background: white;
    border-radius: 10px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.06);
  }}
  .action-bar .action-label {{
    font-size: 0.75rem;
    color: #5a6f8a;
    margin-right: 4px;
  }}

  /* ============================================================
     Buttons
  ============================================================ */
  .btn {{
    display: inline-flex;
    align-items: center;
    gap: 6px;
    border: none;
    border-radius: 8px;
    padding: 7px 18px;
    font-size: 0.78rem;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.18s, transform 0.1s;
    white-space: nowrap;
    letter-spacing: 0.02em;
  }}
  .btn:active {{ transform: scale(0.97); }}

  .btn-primary {{
    background: #2c4a7c;
    color: white;
  }}
  .btn-primary:hover {{ background: #3d6499; }}

  .btn-secondary {{
    background: white;
    color: #2c4a7c;
    border: 1.5px solid #2c4a7c;
  }}
  .btn-secondary:hover {{ background: #eef1f6; }}

  .btn-success {{
    background: #1a7a4a !important;
    color: white !important;
    border-color: #1a7a4a !important;
  }}

  /* btn-copy ใน block header */
  .btn-copy {{
    display: inline-flex;
    align-items: center;
    background: #2c4a7c;
    color: white;
    border: none;
    border-radius: 8px;
    padding: 6px 16px;
    font-size: 0.76rem;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.18s, transform 0.1s;
    white-space: nowrap;
  }}
  .btn-copy:hover   {{ background: #3d6499; }}
  .btn-copy:active  {{ transform: scale(0.97); }}
  .btn-copy.success {{ background: #1a7a4a; }}

  /* ============================================================
     Customer Block
  ============================================================ */
  .customer-block {{
    background: white;
    border-radius: 10px;
    margin-bottom: 16px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.07);
    overflow: hidden;
    border: 1px solid #d8e0ec;
  }}

  .block-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 12px 18px;
    background: #f4f7fc;
    border-bottom: 1px solid #d8e0ec;
    gap: 12px;
  }}
  .block-header-left {{
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
    min-width: 0;
  }}

  /* Checkbox */
  .row-check {{
    width: 16px;
    height: 16px;
    accent-color: #2c4a7c;
    cursor: pointer;
    flex-shrink: 0;
  }}

  .customer-title {{
    font-size: 0.9rem;
    font-weight: 700;
    color: #1e2d45;
    cursor: pointer;
  }}
  .block-meta {{
    font-size: 0.74rem;
    color: #6a7f9a;
  }}

  /* ============================================================
     Table Wrapper — scroll horizontal, กรอบคม
  ============================================================ */
  .table-scroll {{
    overflow-x: auto;        /* scroll ซ้าย-ขวาเมื่อ column เกิน */
    overflow-y: visible;
    padding: 14px 18px 16px;
    max-width: 100%;
  }}

  /* table เอง: ไม่มี CSS style — ปล่อย browser default
     เพื่อให้ paste ใน Outlook ได้ border grid line เดิม      */
  .table-scroll table {{
    border-collapse: collapse;
    min-width: 100%;          /* บังคับให้ scroll แทน wrap */
    white-space: nowrap;      /* ป้องกัน column ตัดบรรทัด */
  }}

  /* ============================================================
     Divider + Footer
  ============================================================ */
  hr {{ border: none; border-top: 1px solid #d8e0ec; margin: 20px 0; }}

  .footer {{
    text-align: center;
    font-size: 0.7rem;
    color: #9aaabb;
    margin-top: 28px;
    padding-bottom: 16px;
  }}

  /* ============================================================
     Toast notification
  ============================================================ */
  #toast {{
    position: fixed;
    bottom: 28px;
    right: 28px;
    background: #1e2d45;
    color: white;
    font-size: 0.8rem;
    font-weight: 600;
    padding: 10px 22px;
    border-radius: 8px;
    box-shadow: 0 4px 16px rgba(0,0,0,0.18);
    opacity: 0;
    transition: opacity 0.25s;
    pointer-events: none;
    z-index: 9999;
  }}
  #toast.show {{ opacity: 1; }}

</style>
</head>
<body>

<!-- ===== Page Header ===== -->
<div class="page-header">
  <div>
    <h1>Credit Overdue Daily Report</h1>
    <div class="meta">
      CompanyCode: {TARGET_COMPANY} &nbsp;|&nbsp;
      Report Date: {report_date_str} &nbsp;|&nbsp;
      {len(customers)} customer(s) with overdue
    </div>
  </div>
  <div class="badge">Confidential</div>
</div>

<!-- ===== Email Preview ===== -->
<div class="email-box">
  <div class="label">Email Message Preview</div>
  <pre id="email-text">{sig_escaped}</pre>
</div>

<hr>

<!-- ===== Action Bar ===== -->
<div class="action-bar">
  <span class="action-label">Bulk Actions:</span>
  <button class="btn btn-primary" onclick="copyAll()">
    Copy All Tables
  </button>
  <button class="btn btn-secondary" onclick="copySelected()">
    Copy Multi (Selected)
  </button>
  <span id="selected-count" style="font-size:0.74rem;color:#6a7f9a;margin-left:6px;">
    0 selected
  </span>
</div>

<!-- ===== Customer Blocks ===== -->
{customer_blocks_html}

<!-- ===== Footer ===== -->
<div class="footer">
  Credit Monitoring System &nbsp;|&nbsp; {report_date_str}
</div>

<!-- ===== Toast ===== -->
<div id="toast"></div>

<script>

  /* ---- Signature color ---- */
  (function () {{
    const pre = document.getElementById('email-text');
    if (!pre) return;
    const lines = pre.textContent.split('\\n');
    let inSig = false;
    pre.innerHTML = lines.map(line => {{
      if (!inSig && (line.startsWith('Best Regards') || line.startsWith('*-'))) inSig = true;
      return inSig ? '<span class="sig">' + line + '</span>' : line;
    }}).join('\\n');
  }})();

  /* ---- Checkbox counter ---- */
  document.querySelectorAll('.row-check').forEach(function(chk) {{
    chk.addEventListener('change', updateCount);
  }});

  function updateCount() {{
    const n = document.querySelectorAll('.row-check:checked').length;
    document.getElementById('selected-count').textContent =
      n === 0 ? '0 selected' : n + ' selected';
  }}

  /* ---- Toast helper ---- */
  function showToast(msg) {{
    const t = document.getElementById('toast');
    t.textContent = msg;
    t.classList.add('show');
    setTimeout(() => t.classList.remove('show'), 2200);
  }}

  /* ---- Core: copy table HTML ใน div เป็น rich HTML ---- */
  async function copyTableHTML(tableWrapperEl, btnEl) {{
    const table = tableWrapperEl.querySelector('table');
    if (!table) return;

    const htmlBlob = new Blob(
      ['<meta charset="utf-8">' + table.outerHTML],
      {{ type: 'text/html' }}
    );
    const textBlob = new Blob([table.innerText], {{ type: 'text/plain' }});

    try {{
      await navigator.clipboard.write([
        new ClipboardItem({{ 'text/html': htmlBlob, 'text/plain': textBlob }})
      ]);
    }} catch (e) {{
      /* fallback execCommand */
      const range = document.createRange();
      range.selectNode(table);
      window.getSelection().removeAllRanges();
      window.getSelection().addRange(range);
      document.execCommand('copy');
      window.getSelection().removeAllRanges();
    }}

    if (btnEl) {{
      btnEl.classList.add('success');
      const orig = btnEl.textContent;
      btnEl.textContent = 'Copied';
      setTimeout(() => {{
        btnEl.textContent = orig;
        btnEl.classList.remove('success');
      }}, 2000);
    }}
  }}

  /* ---- Copy single table ---- */
  function copySingle(sectionId, btn) {{
    const el = document.getElementById(sectionId);
    copyTableHTML(el, btn).then(() => showToast('Table copied'));
  }}

  /* ---- Copy ALL tables — merge เป็น HTML เดียว ---- */
  async function copyAll() {{
    const wrappers = document.querySelectorAll('.table-scroll');
    await copyMerged(wrappers, 'All tables copied (' + wrappers.length + ')');
  }}

  /* ---- Copy SELECTED tables ---- */
  async function copySelected() {{
    const checked = document.querySelectorAll('.row-check:checked');
    if (checked.length === 0) {{
      showToast('No table selected');
      return;
    }}
    const wrappers = Array.from(checked).map(chk => {{
      const idx = chk.getAttribute('data-index');
      return document.getElementById('report_' + idx);
    }}).filter(Boolean);
    await copyMerged(wrappers, checked.length + ' table(s) copied');
  }}

  /* ---- Merge หลาย table เป็น HTML เดียวแล้ว copy ---- */
  async function copyMerged(wrappers, toastMsg) {{
    let combinedHTML = '<meta charset="utf-8">';
    let combinedText = '';

    wrappers.forEach(function(w, i) {{
      const table = w ? w.querySelector('table') : null;
      if (!table) return;
      /* หา customer title จาก block parent */
      const block  = w.closest('.customer-block');
      const title  = block ? block.querySelector('.customer-title') : null;
      const label  = title ? title.textContent.trim() : ('Table ' + (i + 1));

      combinedHTML += '<p><strong>' + label + '</strong></p>';
      combinedHTML += table.outerHTML;
      combinedHTML += '<br>';

      combinedText += label + '\\n';
      combinedText += table.innerText + '\\n\\n';
    }});

    const htmlBlob = new Blob([combinedHTML], {{ type: 'text/html' }});
    const textBlob = new Blob([combinedText], {{ type: 'text/plain' }});

    try {{
      await navigator.clipboard.write([
        new ClipboardItem({{ 'text/html': htmlBlob, 'text/plain': textBlob }})
      ]);
      showToast(toastMsg);
    }} catch (e) {{
      showToast('Copy failed — try selecting manually');
    }}
  }}

</script>
</body>
</html>"""

# ---------------------------------------------------------------------------
# เขียน HTML ลง tempfile แล้วเปิดด้วย webbrowser (เหมือน example)
# ทำงานฝั่ง server — เหมาะกับ local Streamlit
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# เปิด tab ตาม case:
#   has_overdue = True  → เปิด 2 tab (report HTML + Outlook)
#   has_overdue = False → เปิด Outlook อย่างเดียว
# ---------------------------------------------------------------------------
def _open_two_tabs(report_html: str, outlook_url: str, has_overdue: bool):
    if has_overdue:
        # เขียน tempfile + เปิด 2 tab
        tmp = tempfile.NamedTemporaryFile(
            suffix=".html",
            delete=False,
            mode="w",
            encoding="utf-8",
        )
        tmp.write(report_html)
        tmp.close()

        webbrowser.open(f"file://{tmp.name}")
        webbrowser.open(outlook_url)
    else:
        # ไม่มี overdue → เปิด Outlook อย่างเดียว ไม่ต้อง copy table
        webbrowser.open(outlook_url)

# ---------------------------------------------------------------------------
# Outlook URL — body สั้น (ไม่รวม table)
# ---------------------------------------------------------------------------
def _build_outlook_url(email_to: str, subject: str, short_body: str) -> str:
    base  = "https://outlook.office.com/mail/deeplink/compose"
    query = urllib.parse.urlencode(
        {"to": email_to, "subject": subject, "body": short_body},
        quote_via=urllib.parse.quote,
    )
    return f"{base}?{query}"


def _build_short_body(
    has_overdue: bool,
    salutation: str,
    closing: str,
    signature: str,
) -> str:
    lines = [salutation, ""]
    if has_overdue:
        lines += ["มี OverDue ตามรายการด้านล่างค่ะ", "", "[paste table here]"]
    else:
        lines += ["วันนี้ไม่มี OverDue ค่ะ"]
    lines += ["", closing, "", signature]
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# KPI Cards
# ---------------------------------------------------------------------------
def _render_kpi_cards(df: pd.DataFrame):
    total_amount    = float(df["OverdueAmount"].sum()) if "OverdueAmount" in df.columns else 0.0
    total_records   = len(df)
    total_customers = int(df["CustomerName"].nunique()) if "CustomerName" in df.columns else 0

    kpi_data = [
        ("Total Overdue Amount (THB)", f"{total_amount:,.2f}"),
        ("Total Overdue Records",      f"{total_records:,}"),
        ("Affected Customers",         f"{total_customers:,}"),
    ]

    cols = st.columns(3)
    for col, (label, value) in zip(cols, kpi_data):
        with col:
            st.markdown(
                f"""
                <div style="
                    border: 1px solid #c8cdd4;
                    border-radius: 8px;
                    padding: 18px 22px;
                    background: #ffffff;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
                ">
                    <div style="
                        font-size: 0.72rem;
                        font-weight: 600;
                        text-transform: uppercase;
                        letter-spacing: 0.07em;
                        color: #6b7280;
                        margin-bottom: 8px;
                    ">{label}</div>
                    <div style="
                        font-size: 1.45rem;
                        font-weight: 700;
                        color: #111827;
                        letter-spacing: -0.01em;
                        line-height: 1.2;
                    ">{value}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

# ---------------------------------------------------------------------------
# Section 1: ตารางดิบ + Dynamic Filter
# ---------------------------------------------------------------------------
def _render_raw_table_with_filter(df_overdue: pd.DataFrame) -> pd.DataFrame:
    st.markdown(section_header("Raw Overdue Data"), unsafe_allow_html=True)

    with st.form("overdue_daily_filter_form"):
        filter_cols  = st.columns(4)
        user_filters = {}

        for idx, col_name in enumerate(df_overdue.columns):
            with filter_cols[idx % 4]:
                unique_vals = df_overdue[col_name].dropna().unique()
                try:
                    unique_vals = sorted(unique_vals)
                except TypeError:
                    unique_vals = sorted(unique_vals, key=str)
                sel = st.multiselect(col_name, unique_vals, key=f"od_daily_filter_{col_name}")
                if sel:
                    user_filters[col_name] = sel

        _, btn_col = st.columns([5, 1])
        with btn_col:
            st.form_submit_button("Apply", type="primary", use_container_width=True)

    filtered_od = df_overdue.copy()
    for fc, fv in user_filters.items():
        filtered_od = filtered_od[filtered_od[fc].isin(fv)]

    view_mode  = st.radio(
        "Display:", ["Preview (30 rows)", "Show All"],
        horizontal=True, key="od_daily_view_mode",
    )
    display_df = filtered_od.head(30) if "Preview" in view_mode else filtered_od
    st.dataframe(display_df, use_container_width=True, height=400)
    st.caption(
        f"Filtered: {filtered_od.shape[1]} cols x {filtered_od.shape[0]} rows"
        f" (from {df_overdue.shape[0]} total)"
    )
    return filtered_od


# ---------------------------------------------------------------------------
# Section 2: Auto-Filtering + result + Compose + Outlook
# ---------------------------------------------------------------------------
def _render_auto_filter_section(df_source: pd.DataFrame):
    st.divider()
    st.markdown(section_header("Auto-Filtering Result"), unsafe_allow_html=True)

    col_space, col_toggle, col_date, col_btn = st.columns([4, 2, 2, 2])

    with col_toggle:
        use_custom_date = st.toggle(
            "Custom Date",
            value=False,
            key="od_daily_use_custom_date",
            help="Off = วันปัจจุบัน  |  On = เลือกวันเองสำหรับทดสอบ",
        )
    with col_date:
        if use_custom_date:
            target_date = st.date_input(
                "Select Date",
                value=date.today(),
                key="od_daily_custom_date",
                format="YYYY/MM/DD",
                label_visibility="collapsed",
            )
        else:
            target_date = date.today()
            st.markdown(
                f"<div style='padding:6px 0;color:#3a4a60;font-size:0.85rem;'>"
                f"Using today: <b>{target_date.strftime('%d %b %Y')}</b></div>",
                unsafe_allow_html=True,
            )
    with col_btn:
        run_auto = st.button(
            "Auto-Filtering",
            type="primary",
            use_container_width=True,
            key="od_daily_auto_filter_btn",
        )

    st.caption(
        f"Filter: CompanyCode = {TARGET_COMPANY}  |  "
        f"CollectionDate = {target_date.strftime('%Y%m%d')}  |  "
        f"OverdueAmount > 0"
    )

    if run_auto:
        st.session_state["od_daily_auto_filter_ran"] = True

    if not st.session_state.get("od_daily_auto_filter_ran", False):
        st.info("Click 'Auto-Filtering' to apply business rules and see the result.")
        return

    filtered_df = _apply_auto_filter(df_source, target_date)
    today_str   = target_date.strftime("%Y%m%d")
    has_overdue = not filtered_df.empty

    st.caption(
        f"Report Date: {target_date.strftime('%d %B %Y')}  |  "
        f"Source file: {st.session_state.get('latest_overdue_name', '-')}"
    )

    if not has_overdue:
        st.success(
            f"No overdue records — "
            f"CompanyCode: {TARGET_COMPANY}, CollectionDate: {today_str}."
        )
        st.dataframe(pd.DataFrame(columns=df_source.columns), use_container_width=True, height=100)
    else:
        st.divider()
        st.markdown(section_header("Key Metrics"), unsafe_allow_html=True)
        _render_kpi_cards(filtered_df)

        st.divider()
        st.markdown(section_header("Overdue Records by Customer"), unsafe_allow_html=True)

        if "CustomerName" in filtered_df.columns:
            for customer in sorted(filtered_df["CustomerName"].dropna().unique()):
                group = filtered_df[filtered_df["CustomerName"] == customer].reset_index(drop=True)
                total = float(group["OverdueAmount"].sum()) if "OverdueAmount" in group.columns else 0.0
                with st.expander(
                    f"{customer}   |   {len(group)} records   |   OverdueAmount: {total:,.2f} THB",
                    expanded=False,
                ):
                    st.dataframe(group, use_container_width=True)
        else:
            with st.expander("All Overdue Records", expanded=True):
                st.dataframe(filtered_df, use_container_width=True)


        st.divider()
        st.markdown(section_header("Export Report"), unsafe_allow_html=True)

        excel_bytes = _build_excel(filtered_df)
        filename    = f"overdue_daily_{today_str}.xlsx"

        # inject CSS ให้ปุ่ม download_button ตัวนี้เป็นสีเขียว
        st.markdown(
            """
            <style>
            div[data-testid="stDownloadButton"] > button {
                background-color: #1a7a4a;
                color: white;
                border: none;
                font-weight: 600;
            }
            div[data-testid="stDownloadButton"] > button:hover {
                background-color: #15623b;
                color: white;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        st.download_button(
            label="Download .xlsx  (Summary + Sheet per Customer)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ==========================================================================
    # Compose & Send
    # ==========================================================================
    st.divider()
    st.markdown(section_header("Compose & Send via Outlook"), unsafe_allow_html=True)

    auto_subject = _build_subject(target_date)

    cfg_col1, cfg_col2 = st.columns(2)
    with cfg_col1:
        email_to = st.text_input("To", value=DEFAULT_RECIPIENT, key="od_daily_email_to")
    with cfg_col2:
        email_subject = st.text_input("Subject", value=auto_subject, key="od_daily_email_subject")

    with st.expander("Edit Email Content", expanded=True):
        body_salutation = st.text_input(
            "Salutation", value=DEFAULT_SALUTATION, key="od_daily_salutation"
        )
        body_closing = st.text_area(
            "Closing", value=DEFAULT_CLOSING, height=80, key="od_daily_closing"
        )
        body_signature = st.text_area(
            "Signature", value=DEFAULT_SIGNATURE, height=140, key="od_daily_signature"
        )

    short_body = _build_short_body(
        has_overdue = has_overdue,
        salutation  = body_salutation,
        closing     = body_closing,
        signature   = body_signature,
    )

    with st.expander("Preview Email (Outlook body)", expanded=False):
        st.text(short_body)

    _, outlook_col = st.columns([7, 3])
    with outlook_col:
        go_clicked = st.button(
            "Go to Outlook  →",
            type="primary",
            use_container_width=True,
            key="od_daily_go_outlook",
            disabled=not email_to.strip(),
        )

    if go_clicked and email_to.strip():
        outlook_url = _build_outlook_url(
            email_to.strip(), email_subject.strip(), short_body
        )

        # ส่ง has_overdue เข้าไปด้วย — ตัดสินใจใน _open_two_tabs
        report_html = _build_report_html(filtered_df, target_date, short_body)
        _open_two_tabs(report_html, outlook_url, has_overdue)

        if has_overdue:
            st.success(
                "เปิด 2 tab แล้วเน้ออ้าย  \n"
                "**Tab 1:** Report — กด 'Copy Table' ต่อ customer แล้ว paste ใน Outlook  \n"
                "**Tab 2:** Outlook Draft — วาง table หลัง [paste table here]"
            )
        else:
            st.success("ไม่มี Overdue — เปิด Outlook Draft แล้ว กดส่งได้เลย")


# ---------------------------------------------------------------------------
# Main render
# ---------------------------------------------------------------------------
def render():
    st.markdown(
        dash_title_bar(
            "Overdue Daily",
            "Jelly Section — Review data, auto-filter, compose and send via Outlook",
        ),
        unsafe_allow_html=True,
    )

    if not st.session_state.get("data_processed", False):
        st.info("Please run the pipeline in 'Loading and Processing Data' first.")
        return

    df_overdue = st.session_state.get("df_overdue")
    if df_overdue is None or (hasattr(df_overdue, "empty") and df_overdue.empty):
        st.warning("df_overdue is empty. Please re-process the data pipeline.")
        return

    filtered_by_user = _render_raw_table_with_filter(df_overdue)
    _render_auto_filter_section(filtered_by_user)