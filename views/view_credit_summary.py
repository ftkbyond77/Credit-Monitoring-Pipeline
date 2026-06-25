# views/view_credit_summary.py
#
# PNONG Section — Credit Summary
# Sections:
#   1. Raw Overdue Table   (เหมือน view_overdue_daily.py + dynamic filter)
#   2. Merged Table        (Overdue left-join Credit Availability via As-of Join)
#   3. Summary Report      (Auto-filter per OriginalDueDate, expand by company,
#                           customizable toggle, total row per company,
#                           statistics footer, export .xlsx)

import io
from datetime import date, datetime

import pandas as pd
import numpy as np
import streamlit as st

from components import dash_title_bar, section_header


# =============================================================================
# Constants
# =============================================================================

# คอลัมน์ Credit Availability ที่ต้องการ merge (Customer Code → Est. Amount)
AVAIL_COLS_WANTED = [
    "CUSTOMER_CODE",
    "CUSTOMER_NAME",
    "CLEAN_CREDIT_MB",
    "CURRENT_DEBT_MILLION_THB",
    "CURRENT_DEBT_MILLION_THB_PERCENT",
    "EST_FURTHER_AMOUNT",
    "EST_DEBT",
    "ESTIMATE_AMOUNT",
    "DATE",           # SnapshotDate — ใช้สำหรับ As-of Join
    "SOURCE_SHEET",   # ปีที่มาของ snapshot
]

# คอลัมน์ที่แสดงผลหลัง merge (ไม่รวม internal helpers)
MERGED_DISPLAY_COLS = [
    "CompanyCode",
    "Customer",
    "CustomerName",
    "InvoiceDocument",
    "OriginalDueDate",
    "CollectionDate",
    "OverdueAmount",
    "InvoiceAmount",
    "CLEAN_CREDIT_MB",
    "CURRENT_DEBT_MILLION_THB",
    "CURRENT_DEBT_MILLION_THB_PERCENT",
    "ESTIMATE_AMOUNT",
    "AVAIL_SNAPSHOT_DATE",   # วันที่ snapshot ที่ถูกเลือก (debug / transparency)
]


# =============================================================================
# Date Utilities
# =============================================================================

def _parse_overdue_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def _parse_avail_date(series: pd.Series) -> pd.Series:
    """
    Convert DATE column จาก df_avail → pd.Timestamp

    et_pipeline.py ได้ parse + realign ไปแล้วใน _transform_availability()
    output คือ datetime.date object (Python builtin) หรือ np.nan

    ดังนั้นที่นี่แค่ pd.to_datetime() ตรงๆ ได้เลย
    ไม่ต้อง dayfirst / format เพราะ date object ไม่มี ambiguity แล้ว
    """
    return pd.to_datetime(series, errors="coerce")


# =============================================================================
# As-of Join Logic
# =============================================================================

def _as_of_join_availability(
    df_overdue: pd.DataFrame,
    df_avail: pd.DataFrame,
) -> pd.DataFrame:
    """
    As-of Join: เลือก snapshot ล่าสุดที่ AvailDate <= OriginalDueDate ต่อ invoice

    Rules:
      - Overdue นำ (left join) — ทุก row ถูกเก็บไว้เสมอ
      - Row ที่ OriginalDueDate = NaT → ไม่สามารถ as-of ได้ → avail cols = NaN
      - Row ที่ไม่มี customer ใน avail → avail cols = NaN
      - dtype ทั้งสองฝั่ง normalize เป็น datetime64[us] ก่อน merge_asof
        (avail DATE มาจาก et_pipeline เป็น datetime.date → concat → datetime64[s]
         overdue OriginalDueDate → pd.to_datetime → datetime64[us]
         merge_asof pandas >= 2.0 บังคับ exact same dtype)
    """
    if df_avail is None or df_avail.empty:
        for col in AVAIL_COLS_WANTED:
            if col not in ("CUSTOMER_CODE", "DATE", "SOURCE_SHEET"):
                df_overdue[col] = np.nan
        df_overdue["AVAIL_SNAPSHOT_DATE"] = pd.NaT
        return df_overdue

    # ------------------------------------------------------------------
    # 1. Prepare overdue
    #    normalize OriginalDueDate → datetime64[us]
    #    แยก NaT rows ออกก่อน — merge_asof ห้าม null บน left key
    # ------------------------------------------------------------------
    df_od = df_overdue.copy()

    if "OriginalDueDate" in df_od.columns:
        df_od["OriginalDueDate"] = (
            pd.to_datetime(df_od["OriginalDueDate"], errors="coerce")
            .astype("datetime64[us]")
        )
    else:
        df_od["OriginalDueDate"] = pd.NaT

    # Join key: Customer (int)
    if "Customer" in df_od.columns:
        df_od["_join_key"] = (
            pd.to_numeric(df_od["Customer"], errors="coerce")
            .fillna(0).astype(int)
        )
    else:
        cand = [
            c for c in df_od.columns
            if "customer" in c.lower() and "name" not in c.lower()
        ]
        df_od["_join_key"] = (
            pd.to_numeric(df_od[cand[0]], errors="coerce")
            .fillna(0).astype(int)
            if cand else 0
        )

    # เก็บ original positional index ไว้สำหรับ reconstruct ลำดับ
    df_od = df_od.reset_index(drop=True)
    df_od["_orig_idx"] = df_od.index

    # แยก: row ที่ merge_asof ได้ vs row ที่ NaT (ต้องข้ามการ merge)
    nat_mask      = df_od["OriginalDueDate"].isna()
    df_od_valid   = df_od[~nat_mask].copy()   # มี date จริง → merge ได้
    df_od_nat     = df_od[nat_mask].copy()    # NaT → ข้ามไป, avail = NaN

    # ------------------------------------------------------------------
    # 2. Prepare avail
    #    DATE จาก et_pipeline เป็น datetime.date object
    #    pd.concat หลาย sheet → datetime64[s]
    #    normalize → datetime64[us] ให้ตรงกับฝั่ง overdue
    # ------------------------------------------------------------------
    avail_exist_cols = [c for c in AVAIL_COLS_WANTED if c in df_avail.columns]
    df_av = df_avail[avail_exist_cols].copy()

    if "DATE" in df_av.columns:
        df_av["DATE"] = (
            pd.to_datetime(df_av["DATE"], errors="coerce")
            .astype("datetime64[us]")
        )
    else:
        df_av["DATE"] = pd.NaT

    df_av["CUSTOMER_CODE"] = (
        pd.to_numeric(df_av["CUSTOMER_CODE"], errors="coerce")
        .fillna(0).astype(int)
    )

    # Drop avail rows ที่ DATE เป็น NaT (as-of ไม่ได้)
    df_av = df_av.dropna(subset=["DATE"]).copy()

    # Dedup ต่อ (CUSTOMER_CODE, DATE): หลาย source sheet วันเดียวกัน → เลือกล่าสุด
    df_av = (
        df_av
        .sort_values(["CUSTOMER_CODE", "DATE", "SOURCE_SHEET"], ascending=True)
        .drop_duplicates(subset=["CUSTOMER_CODE", "DATE"], keep="last")
        .reset_index(drop=True)
    )

    avail_value_cols = [
        c for c in avail_exist_cols
        if c not in ("CUSTOMER_CODE", "DATE", "SOURCE_SHEET")
    ]

    # ------------------------------------------------------------------
    # 3. เตรียม NaT rows — ใส่ NaN ให้ avail cols ทุกตัว
    # ------------------------------------------------------------------
    for col in avail_value_cols:
        df_od_nat[col] = np.nan
    df_od_nat["AVAIL_SNAPSHOT_DATE"] = pd.NaT

    # ------------------------------------------------------------------
    # 4. As-of Join — loop per customer (เฉพาะ valid rows)
    # ------------------------------------------------------------------
    result_rows = []

    if not df_od_valid.empty:
        unique_customers = df_od_valid["_join_key"].unique()

        for cust_code in unique_customers:
            od_sub = df_od_valid[df_od_valid["_join_key"] == cust_code].copy()
            av_sub = df_av[df_av["CUSTOMER_CODE"] == cust_code].copy()

            # sort by OriginalDueDate (required by merge_asof)
            od_sub = od_sub.sort_values("OriginalDueDate").reset_index(drop=True)

            if av_sub.empty:
                # customer ไม่มีใน avail → ใส่ NaN
                for col in avail_value_cols:
                    od_sub[col] = np.nan
                od_sub["AVAIL_SNAPSHOT_DATE"] = pd.NaT
                result_rows.append(od_sub)
                continue

            av_sub = av_sub.sort_values("DATE").reset_index(drop=True)

            # merge_asof: DATE <= OriginalDueDate → direction="backward"
            # ทั้งสองฝั่ง datetime64[us] → ไม่เกิด MergeError / ValueError
            merged = pd.merge_asof(
                od_sub,
                av_sub[["DATE"] + avail_value_cols].rename(
                    columns={"DATE": "_avail_date"}
                ),
                left_on="OriginalDueDate",
                right_on="_avail_date",
                direction="backward",
            )

            merged["AVAIL_SNAPSHOT_DATE"] = merged["_avail_date"]
            merged = merged.drop(columns=["_avail_date"])
            result_rows.append(merged)

    # ------------------------------------------------------------------
    # 5. Reconstruct: รวม valid rows + NaT rows → sort ตาม _orig_idx
    # ------------------------------------------------------------------
    parts_to_concat = []

    if result_rows:
        parts_to_concat.append(pd.concat(result_rows, ignore_index=True))

    if not df_od_nat.empty:
        parts_to_concat.append(df_od_nat)

    if not parts_to_concat:
        # กรณีไม่มีข้อมูลเลย
        for col in avail_value_cols:
            df_od[col] = np.nan
        df_od["AVAIL_SNAPSHOT_DATE"] = pd.NaT
        return df_od.drop(columns=["_join_key", "_orig_idx"], errors="ignore")

    result_df = (
        pd.concat(parts_to_concat, ignore_index=True)
        .sort_values("_orig_idx")
        .drop(columns=["_join_key", "_orig_idx"], errors="ignore")
        .reset_index(drop=True)
    )

    return result_df

# =============================================================================
# Excel Export
# =============================================================================

def _build_summary_excel(
    df_overdue: pd.DataFrame,
    df_merged: pd.DataFrame,
    summary_data: dict,   # {company_name: {"df": ..., "total": float}}
    stats_df: pd.DataFrame,
) -> bytes:
    """
    สร้าง .xlsx แบ่ง sheet:
      - Sheet 1 : Summary          (ข้อมูล overdue filtered ทั้งหมด รวมทุกบริษัท)
      - Sheet 2 : Total_Summary    (total OverdueAmount ต่อบริษัท + เว้น 2 แถวว่าง)
      - Sheet 3+: per company      (1 sheet ต่อ 1 company พร้อม Total row)
      - Sheet สุดท้าย: Statistics
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ------------------------------------------------------------------
    # Style tokens
    # ------------------------------------------------------------------
    HEADER_FILL  = PatternFill(start_color="1A7A4A", end_color="1A7A4A", fill_type="solid")
    TOTAL_FILL   = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    STATS_FILL   = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    TOTAL_FONT   = Font(bold=True, size=10, color="1A7A4A")
    STATS_FONT   = Font(color="FFFFFF", bold=True, size=10)
    THIN_SIDE    = Side(style="thin", color="C8CDD4")
    CELL_BORDER  = Border(
        left=THIN_SIDE, right=THIN_SIDE,
        top=THIN_SIDE,  bottom=THIN_SIDE,
    )
    LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
    CENTER_ALIGN = Alignment(horizontal="center",  vertical="center")
    RIGHT_ALIGN  = Alignment(horizontal="right",   vertical="center")

    # ------------------------------------------------------------------
    # Helper functions
    # ------------------------------------------------------------------
    def _style_header_row(ws, fill=HEADER_FILL, font=HEADER_FONT):
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = CENTER_ALIGN
            cell.border    = CELL_BORDER

    def _style_data_rows(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.border    = CELL_BORDER
                cell.alignment = LEFT_ALIGN

    def _auto_col_width(ws):
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    def _freeze_and_filter(ws):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------
    # Build workbook
    # ------------------------------------------------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # --------------------------------------------------------------
        # Sheet 1: Summary — ข้อมูล overdue filtered รวมทุกบริษัท
        # --------------------------------------------------------------
        df_overdue.to_excel(writer, sheet_name="Summary", index=False)
        ws_s1 = writer.sheets["Summary"]
        _style_header_row(ws_s1)
        _style_data_rows(ws_s1)
        _auto_col_width(ws_s1)
        _freeze_and_filter(ws_s1)

        # --------------------------------------------------------------
        # Sheet 2: Total_Summary
        # แสดง total OverdueAmount ต่อบริษัท
        # เว้น 2 แถวว่างหลังแต่ละบริษัทเพื่อแยกวิเคราะห์ต่อได้
        # --------------------------------------------------------------
        ws_total = writer.book.create_sheet("Total_Summary")

        if summary_data and any(
            "OverdueAmount" in info["df"].columns
            for info in summary_data.values()
        ):
            # Header
            ts_headers = ["CustomerName", "OverdueAmount"]
            for col_idx, h in enumerate(ts_headers, start=1):
                cell            = ws_total.cell(row=1, column=col_idx, value=h)
                cell.fill       = HEADER_FILL
                cell.font       = HEADER_FONT
                cell.alignment  = CENTER_ALIGN
                cell.border     = CELL_BORDER

            current_row = 2

            for company_name, info in summary_data.items():
                total_overdue = info["total"]   # คำนวณมาแล้วจาก _render_section3_summary

                # Data row: CustomerName | total OverdueAmount
                cell_name            = ws_total.cell(row=current_row, column=1, value=company_name)
                cell_name.border     = CELL_BORDER
                cell_name.alignment  = LEFT_ALIGN

                cell_amt             = ws_total.cell(row=current_row, column=2, value=total_overdue)
                cell_amt.fill        = TOTAL_FILL
                cell_amt.font        = TOTAL_FONT
                cell_amt.border      = CELL_BORDER
                cell_amt.alignment   = RIGHT_ALIGN
                cell_amt.number_format = "#,##0.00"

                # เว้น 2 แถวว่างหลังแต่ละบริษัท
                current_row += 3   # 1 data row + 2 blank rows

            ws_total.column_dimensions["A"].width = 45
            ws_total.column_dimensions["B"].width = 22
            ws_total.freeze_panes = "A2"

        else:
            ws_total.cell(
                row=1, column=1,
                value="OverdueAmount or CustomerName column not found",
            )

        # --------------------------------------------------------------
        # Sheet 3+: per company — 1 sheet ต่อ 1 company + Total row
        # --------------------------------------------------------------
        for company_name, info in summary_data.items():
            df_co   = info["df"]
            total   = info["total"]

            safe_name = (
                str(company_name)[:28] + "..."
                if len(str(company_name)) > 28
                else str(company_name)
            )

            # เพิ่ม Total row ท้าย
            total_row = {col: "" for col in df_co.columns}
            if "CustomerName" in df_co.columns:
                total_row["CustomerName"] = "TOTAL"
            if "OverdueAmount" in df_co.columns:
                total_row["OverdueAmount"] = total

            df_export = pd.concat(
                [df_co, pd.DataFrame([total_row])],
                ignore_index=True,
            )
            df_export.to_excel(writer, sheet_name=safe_name, index=False)

            ws_co = writer.sheets[safe_name]
            _style_header_row(ws_co)
            _style_data_rows(ws_co)

            # Style Total row (แถวสุดท้าย)
            last_row = ws_co.max_row
            for cell in ws_co[last_row]:
                cell.fill      = TOTAL_FILL
                cell.font      = TOTAL_FONT
                cell.border    = CELL_BORDER
                cell.alignment = LEFT_ALIGN

            _auto_col_width(ws_co)
            _freeze_and_filter(ws_co)

        # --------------------------------------------------------------
        # Sheet สุดท้าย: Statistics
        # --------------------------------------------------------------
        if not stats_df.empty:
            stats_df.to_excel(writer, sheet_name="Statistics", index=False)
            ws_st = writer.sheets["Statistics"]
            _style_header_row(ws_st, fill=STATS_FILL, font=STATS_FONT)
            _style_data_rows(ws_st)
            _auto_col_width(ws_st)

    return output.getvalue()

# =============================================================================
# Section 1 — Raw Overdue Table with Dynamic Filter
# =============================================================================

def _render_section1_raw_table(df_overdue: pd.DataFrame) -> pd.DataFrame:
    """
    แสดงตาราง Credit Overdue พร้อม dynamic filter เหมือน view_overdue_daily.py
    Return: df หลัง filter ที่ user เลือก
    """
    st.markdown(section_header("Section 1 — Credit Overdue Data"), unsafe_allow_html=True)

    with st.form("cs_s1_filter_form"):
        filter_cols  = st.columns(4)
        user_filters = {}

        for idx, col_name in enumerate(df_overdue.columns):
            with filter_cols[idx % 4]:
                unique_vals = df_overdue[col_name].dropna().unique()
                try:
                    unique_vals = sorted(unique_vals)
                except TypeError:
                    unique_vals = sorted(unique_vals, key=str)

                sel = st.multiselect(
                    col_name,
                    unique_vals,
                    key=f"cs_s1_filter_{col_name}",
                )
                if sel:
                    user_filters[col_name] = sel

        _, btn_col = st.columns([5, 1])
        with btn_col:
            st.form_submit_button("Apply Filter", type="primary", use_container_width=True)

    filtered = df_overdue.copy()
    for fc, fv in user_filters.items():
        filtered = filtered[filtered[fc].isin(fv)]

    view_mode  = st.radio(
        "Display:",
        ["Preview (30 rows)", "Show All"],
        horizontal=True,
        key="cs_s1_view_mode",
    )
    display_df = filtered.head(30) if "Preview" in view_mode else filtered

    st.dataframe(display_df, use_container_width=True, height=400)
    st.caption(
        f"Filtered: {filtered.shape[1]} cols x {filtered.shape[0]:,} rows"
        f" (from {df_overdue.shape[0]:,} total)"
    )
    return filtered


# =============================================================================
# Section 2 — Merged Table: Overdue + Credit Availability (As-of Join)
# =============================================================================

def _render_section2_merged_table(
    df_overdue_filtered: pd.DataFrame,
    df_avail: pd.DataFrame,
):
    st.divider()
    st.markdown(
        section_header("Section 2 — Merged: Overdue + Credit Availability (As-of Join)"),
        unsafe_allow_html=True,
    )
    st.caption(
        "Credit Availability ที่ match ใช้หลัก As-of Join: "
        "เลือก snapshot ล่าสุดที่ SnapshotDate <= OriginalDueDate ต่อ invoice"
    )

    # Warn หาก df_avail ไม่มี
    if df_avail is None or (hasattr(df_avail, "empty") and df_avail.empty):
        st.warning("ไม่พบข้อมูล Credit Availability — แสดงเฉพาะ Overdue")
        df_merged = df_overdue_filtered.copy()
    else:
        with st.spinner("Running As-of Join..."):
            # Ensure OriginalDueDate ถูก parse เป็น datetime ก่อน as-of join
            df_od_prep = df_overdue_filtered.copy()
            if "OriginalDueDate" in df_od_prep.columns:
                df_od_prep["OriginalDueDate"] = pd.to_datetime(
                    df_od_prep["OriginalDueDate"], errors="coerce"
                )
            df_merged = _as_of_join_availability(df_od_prep, df_avail)

    # ------------------------------------------------------------------
    # Filter: ปี / เดือน จาก OriginalDueDate + between date
    # ------------------------------------------------------------------
    with st.expander("Filter by OriginalDueDate (Year / Month / Range)", expanded=False):
        f_col1, f_col2, f_col3, f_col4 = st.columns(4)

        # --- available years and months ---
        years_avail  = []
        months_avail = []
        if "OriginalDueDate" in df_merged.columns:
            dt_col = pd.to_datetime(df_merged["OriginalDueDate"], errors="coerce")
            years_avail  = sorted(dt_col.dt.year.dropna().unique().astype(int).tolist())
            months_avail = sorted(dt_col.dt.month.dropna().unique().astype(int).tolist())

        MONTH_LABEL = {
            1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"May", 6:"Jun",
            7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec",
        }

        with f_col1:
            st.markdown("Year (OriginalDueDate)", unsafe_allow_html=True)
            sel_years = st.multiselect(
                "Year",
                options=years_avail,
                default=[],
                key="cs_s2_year",
                label_visibility="collapsed",
            )
        with f_col2:
            st.markdown("Month (OriginalDueDate)", unsafe_allow_html=True)
            month_options = [f"{m} - {MONTH_LABEL[m]}" for m in months_avail]
            sel_month_labels = st.multiselect(
                "Month",
                options=month_options,
                default=[],
                key="cs_s2_month",
                label_visibility="collapsed",
            )
            sel_months = [int(m.split(" - ")[0]) for m in sel_month_labels]

        with f_col3:
            st.markdown("Due Date From", unsafe_allow_html=True)
            s2_date_from = st.date_input(
                "Due From",
                value=None,
                key="cs_s2_date_from",
                format="YYYY/MM/DD",
                label_visibility="collapsed",
            )
        with f_col4:
            st.markdown("Due Date To", unsafe_allow_html=True)
            s2_date_to = st.date_input(
                "Due To",
                value=None,
                key="cs_s2_date_to",
                format="YYYY/MM/DD",
                label_visibility="collapsed",
            )

    # Apply date filters
    df_s2_display = df_merged.copy()
    if "OriginalDueDate" in df_s2_display.columns:
        dt_series = pd.to_datetime(df_s2_display["OriginalDueDate"], errors="coerce")

        if sel_years:
            df_s2_display = df_s2_display[dt_series.dt.year.isin(sel_years)]
            dt_series = dt_series[df_s2_display.index]

        if sel_months:
            df_s2_display = df_s2_display[
                pd.to_datetime(df_s2_display["OriginalDueDate"], errors="coerce")
                .dt.month.isin(sel_months)
            ]

        if s2_date_from:
            df_s2_display = df_s2_display[
                pd.to_datetime(df_s2_display["OriginalDueDate"], errors="coerce")
                >= pd.Timestamp(s2_date_from)
            ]
        if s2_date_to:
            df_s2_display = df_s2_display[
                pd.to_datetime(df_s2_display["OriginalDueDate"], errors="coerce")
                <= pd.Timestamp(s2_date_to)
            ]

    df_s2_display = df_s2_display.reset_index(drop=True)

    # เรียง display columns
    display_cols = [c for c in MERGED_DISPLAY_COLS if c in df_s2_display.columns]
    extra_cols   = [c for c in df_s2_display.columns if c not in display_cols]
    df_s2_show   = df_s2_display[display_cols + extra_cols]

    s2_view = st.radio(
        "Display:",
        ["Preview (30 rows)", "Show All"],
        horizontal=True,
        key="cs_s2_view_mode",
    )
    st.dataframe(
        df_s2_show.head(30) if "Preview" in s2_view else df_s2_show,
        use_container_width=True,
        height=420,
    )
    st.caption(
        f"Merged result: {df_s2_show.shape[1]} cols x {df_s2_show.shape[0]:,} rows"
    )

    return df_merged


# =============================================================================
# Section 3 — Summary Report per Company
# =============================================================================

def _render_section3_summary(df_overdue: pd.DataFrame):
    """
    Summary Report:
      - Auto-filter by OriginalDueDate (default: current month)
      - filter OverdueAmount != 0
      - Toggle สำหรับ customize filter (OriginalDueDate + OverdueAmount multiselect)
      - Expand table รายบริษัท พร้อม Total row
      - Statistics footer
      - Export .xlsx (สีเขียว)
    """
    st.divider()
    st.markdown(section_header("Section 3 — Summary Report"), unsafe_allow_html=True)

    # ------------------------------------------------------------------
    # Prepare: ensure OriginalDueDate เป็น datetime
    # ------------------------------------------------------------------
    df = df_overdue.copy()

    if "OriginalDueDate" in df.columns:
        df["OriginalDueDate"] = pd.to_datetime(
            df["OriginalDueDate"], errors="coerce"
        )
    if "OverdueAmount" in df.columns:
        df["OverdueAmount"] = pd.to_numeric(df["OverdueAmount"], errors="coerce").fillna(0.0)

    # ------------------------------------------------------------------
    # Auto-filter: OriginalDueDate monthly (default = current month)
    # ------------------------------------------------------------------
    today         = date.today()
    default_year  = today.year
    default_month = today.month

    MONTH_LABEL = {
        1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"May", 6:"Jun",
        7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec",
    }

    st.markdown(
        "**Auto-filter:** OriginalDueDate (Monthly, default = current month) "
        "และ OverdueAmount != 0",
        unsafe_allow_html=False,
    )

    af_col1, af_col2, af_col3 = st.columns([2, 2, 2])
    with af_col1:
        af_year = st.selectbox(
            "Year",
            options=sorted(
                df["OriginalDueDate"].dt.year.dropna().unique().astype(int).tolist(),
                reverse=True,
            ) if "OriginalDueDate" in df.columns else [default_year],
            index=0,
            key="cs_s3_af_year",
        )
    with af_col2:
        af_month = st.selectbox(
            "Month",
            options=list(range(1, 13)),
            index=default_month - 1,
            format_func=lambda m: f"{m:02d} - {MONTH_LABEL[m]}",
            key="cs_s3_af_month",
        )
    with af_col3:
        st.markdown("&nbsp;", unsafe_allow_html=True)   # spacer

    # ------------------------------------------------------------------
    # Toggle: Customize Auto-filter
    # ------------------------------------------------------------------
    use_custom = st.toggle(
        "Customize Auto-filter",
        value=False,
        key="cs_s3_custom_toggle",
        help="เปิดเพื่อเลือก OriginalDueDate และ OverdueAmount เพิ่มเติม",
    )

    custom_due_dates    = []
    custom_ovd_amounts  = []

    if use_custom:
        # หา unique values จาก df ที่ผ่าน year/month filter แล้ว
        df_pre = df[
            (df["OriginalDueDate"].dt.year  == af_year)
            & (df["OriginalDueDate"].dt.month == af_month)
            & (df["OverdueAmount"] != 0)
        ]

        cust_col1, cust_col2 = st.columns(2)

        with cust_col1:
            # OriginalDueDate multiselect
            unique_dates = sorted(
                df_pre["OriginalDueDate"].dropna().dt.date.unique().tolist()
            )
            custom_due_dates = st.multiselect(
                "OriginalDueDate (select or leave blank = all)",
                options=unique_dates,
                default=[],
                key="cs_s3_custom_dates",
            )

        with cust_col2:
            # OverdueAmount multiselect (แสดงค่าจริง ไม่ใช่ range)
            unique_amounts = sorted(
                df_pre["OverdueAmount"].dropna().unique().tolist()
            )
            # หากมีมากกว่า 200 ค่า → แสดง "select all" toggle แทน
            if len(unique_amounts) <= 200:
                custom_ovd_amounts = st.multiselect(
                    "OverdueAmount (select or leave blank = all non-zero)",
                    options=unique_amounts,
                    default=[],
                    key="cs_s3_custom_amounts",
                    format_func=lambda x: f"{x:,.2f}",
                )
            else:
                st.info(
                    f"OverdueAmount มี {len(unique_amounts):,} unique values — "
                    "แสดง select-all (ใช้ค่า != 0 ทั้งหมด)"
                )

    # ------------------------------------------------------------------
    # Apply filters
    # ------------------------------------------------------------------
    df_filtered = df.copy()

    # 1. Year + Month filter
    if "OriginalDueDate" in df_filtered.columns:
        df_filtered = df_filtered[
            (df_filtered["OriginalDueDate"].dt.year  == af_year)
            & (df_filtered["OriginalDueDate"].dt.month == af_month)
        ]

    # 2. OverdueAmount != 0
    if "OverdueAmount" in df_filtered.columns:
        df_filtered = df_filtered[df_filtered["OverdueAmount"] != 0]

    # 3. Custom date filter (multiselect)
    if use_custom and custom_due_dates:
        df_filtered = df_filtered[
            df_filtered["OriginalDueDate"].dt.date.isin(custom_due_dates)
        ]

    # 4. Custom amount filter (multiselect)
    if use_custom and custom_ovd_amounts:
        df_filtered = df_filtered[
            df_filtered["OverdueAmount"].isin(custom_ovd_amounts)
        ]

    df_filtered = df_filtered.reset_index(drop=True)

    st.caption(
        f"Filter: Year={af_year}, Month={af_month:02d}-{MONTH_LABEL[af_month]}, "
        f"OverdueAmount != 0  |  Result: {len(df_filtered):,} rows"
    )

    if df_filtered.empty:
        st.info("ไม่พบข้อมูลตามเงื่อนไข Auto-filter")
        return

    # ------------------------------------------------------------------
    # Expand table รายบริษัท (CustomerName)
    # แต่ละบริษัทมี Total row ท้ายสุด
    # ------------------------------------------------------------------
    group_col = "CustomerName" if "CustomerName" in df_filtered.columns else None

    # เก็บไว้สำหรับ export
    summary_company_data = {}

    if group_col:
        companies = sorted(df_filtered[group_col].dropna().unique().tolist())
    else:
        companies = ["All Records"]


    for company in companies:
        if group_col:
            co_df = df_filtered[df_filtered[group_col] == company].reset_index(drop=True)
        else:
            co_df = df_filtered.reset_index(drop=True)

        total_overdue = float(co_df["OverdueAmount"].sum()) \
                        if "OverdueAmount" in co_df.columns else 0.0

        summary_company_data[company] = {"df": co_df, "total": total_overdue}

        with st.expander(
            f"{company}   |   {len(co_df):,} records   |   "
            f"OverdueAmount Total: {total_overdue:,.2f} THB",
            expanded=False,
        ):

            total_row = {}
            for col in co_df.columns:
                if col == "OverdueAmount":
                    total_row[col] = total_overdue
                elif col == group_col:
                    total_row[col] = "TOTAL"
                elif pd.api.types.is_numeric_dtype(co_df[col]):
                    total_row[col] = np.nan
                elif pd.api.types.is_datetime64_any_dtype(co_df[col]):
                    total_row[col] = pd.NaT   # datetime → NaT ไม่ใช่ ""
                else:
                    total_row[col] = ""       # object/str เท่านั้น

            df_display = pd.concat(
                [co_df, pd.DataFrame([total_row])],
                ignore_index=True,
            )

            # ----------------------------------------------------------------
            # Highlight แถว TOTAL
            # ----------------------------------------------------------------
            def _highlight_total(row):
                name_val = (
                    str(row.get(group_col, ""))
                    if group_col
                    else str(row.get("CustomerName", ""))
                )
                if name_val == "TOTAL":
                    return ["background-color: #D9EAD3; font-weight: bold"] * len(row)
                return [""] * len(row)

            st.dataframe(
                df_display.style.apply(_highlight_total, axis=1),
                use_container_width=True,
                height=min(400, (len(df_display) + 2) * 35 + 50),
            )

    # ------------------------------------------------------------------
    # Section 3 Final — Statistics + Export
    # ------------------------------------------------------------------
    st.divider()
    st.markdown(
        section_header("Section 3 Final — Overall Statistics & Export"),
        unsafe_allow_html=True,
    )

    if "OverdueAmount" in df_filtered.columns:
        od_amounts = df_filtered["OverdueAmount"]

        total_sum       = float(od_amounts.sum())
        total_records   = int(len(df_filtered))
        total_companies = int(df_filtered[group_col].nunique()) if group_col else 1
        mean_val        = float(od_amounts.mean())
        median_val      = float(od_amounts.median())
        std_val         = float(od_amounts.std())
        max_val         = float(od_amounts.max())
        min_nonzero     = float(od_amounts[od_amounts > 0].min()) \
                          if (od_amounts > 0).any() else 0.0

        # Per-company aggregate (สำหรับ stats ระดับ company)
        if group_col:
            co_agg = (
                df_filtered.groupby(group_col)["OverdueAmount"]
                .sum().reset_index()
                .rename(columns={"OverdueAmount": "CompanyTotal"})
            )
            avg_per_company = float(co_agg["CompanyTotal"].mean())
            max_company     = co_agg.loc[co_agg["CompanyTotal"].idxmax(), group_col] \
                              if not co_agg.empty else "N/A"
        else:
            avg_per_company = total_sum
            max_company     = "N/A"

        # KPI cards — 4 per row
        stat_rows = [
            [
                ("Total Overdue Amount (THB)",  f"{total_sum:,.2f}",        "Sum of all OverdueAmount",             "danger"),
                ("Total Records",               f"{total_records:,}",        "Rows after filter",                   "info"),
                ("Companies",                   f"{total_companies:,}",      "Unique CustomerName",                  "info"),
                ("Avg per Company (THB)",        f"{avg_per_company:,.2f}",  "Total / Companies",                    "warning"),
            ],
            [
                ("Mean OverdueAmount (THB)",     f"{mean_val:,.2f}",          "Mean per invoice row",                "info"),
                ("Median OverdueAmount (THB)",   f"{median_val:,.2f}",        "Median per invoice row",              "info"),
                ("Std Dev (THB)",                f"{std_val:,.2f}",           "Standard deviation",                  "info"),
                ("Max OverdueAmount (THB)",      f"{max_val:,.2f}",           f"Highest single overdue",             "danger"),
            ],
            [
                ("Min (Non-zero) (THB)",         f"{min_nonzero:,.2f}",       "Smallest positive overdue",           "safe"),
                ("Largest Overdue Company",      str(max_company),            "Company with highest total overdue",  "danger"),
            ],
        ]

        for row_group in stat_rows:
            cols = st.columns(len(row_group))
            for col_ui, (label, value, sub, variant) in zip(cols, row_group):
                variant_map = {
                    "danger":  ("rgba(215,38,61,0.07)",  "#A01F2D"),
                    "warning": ("rgba(181,98,10,0.07)",  "#B5620A"),
                    "safe":    ("rgba(26,122,74,0.07)",  "#1A7A4A"),
                    "info":    ("rgba(27,79,138,0.07)",  "#1B4F8A"),
                }
                bg, accent = variant_map.get(variant, variant_map["info"])
                card_html = (
                    f"<div style='background:{bg};border-left:3px solid {accent};"
                    f"border-radius:6px;padding:10px 14px;margin:4px 0;'>"
                    f"<div style='font-size:0.7rem;color:#3a4a60;font-weight:600;"
                    f"letter-spacing:0.03em;'>{label}</div>"
                    f"<div style='font-size:1.1rem;font-weight:700;color:{accent};"
                    f"margin:2px 0;'>{value}</div>"
                    f"<div style='font-size:0.68rem;color:#7a8a9a;'>{sub}</div>"
                    f"</div>"
                )
                col_ui.markdown(card_html, unsafe_allow_html=True)

        # Build stats df สำหรับ export
        stats_records = [
            {"Metric": "Total Overdue Amount (THB)",   "Value": total_sum},
            {"Metric": "Total Records",                "Value": total_records},
            {"Metric": "Total Companies",              "Value": total_companies},
            {"Metric": "Avg Overdue per Company (THB)","Value": avg_per_company},
            {"Metric": "Mean OverdueAmount (THB)",     "Value": mean_val},
            {"Metric": "Median OverdueAmount (THB)",   "Value": median_val},
            {"Metric": "Std Dev (THB)",                "Value": std_val},
            {"Metric": "Max OverdueAmount (THB)",      "Value": max_val},
            {"Metric": "Min Non-zero (THB)",           "Value": min_nonzero},
            {"Metric": "Largest Overdue Company",      "Value": max_company},
            {"Metric": "Filter Year",                  "Value": af_year},
            {"Metric": "Filter Month",                 "Value": af_month},
            {"Metric": "Export Timestamp",             "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        ]
        stats_df = pd.DataFrame(stats_records)

    else:
        stats_df = pd.DataFrame()

    # ------------------------------------------------------------------
    # Export button (สีเขียว)
    # ------------------------------------------------------------------
    st.markdown("", unsafe_allow_html=True)

    df_avail_state = st.session_state.get("df_avail")
    df_od_state    = st.session_state.get("df_overdue", pd.DataFrame())

    # Merged df สำหรับ export (ใช้ df_filtered เป็น base)
    if df_avail_state is not None and not (
        hasattr(df_avail_state, "empty") and df_avail_state.empty
    ):
        df_od_prep = df_filtered.copy()
        if "OriginalDueDate" in df_od_prep.columns:
            df_od_prep["OriginalDueDate"] = pd.to_datetime(
                df_od_prep["OriginalDueDate"], errors="coerce"
            )
        df_merged_export = _as_of_join_availability(df_od_prep, df_avail_state)
    else:
        df_merged_export = df_filtered.copy()

    export_filename = (
        f"credit_summary_{af_year}{af_month:02d}_"
        f"{datetime.now().strftime('%H%M%S')}.xlsx"
    )

    excel_bytes = _build_summary_excel(
        df_overdue   = df_filtered,
        df_merged    = df_merged_export,
        summary_data = summary_company_data,
        stats_df     = stats_df,
    )

    # inject CSS ให้ download button สีเขียว
    st.markdown(
        """
        <style>
        div[data-testid="stDownloadButton"] > button {
            background-color: #1A7A4A !important;
            color: white !important;
            border: none !important;
            font-weight: 600 !important;
        }
        div[data-testid="stDownloadButton"] > button:hover {
            background-color: #145e38 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.download_button(
        label     = "Export .xlsx  (Overdue Raw | Merged | Summary per Company | Statistics)",
        data      = excel_bytes,
        file_name = export_filename,
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key       = "cs_s3_export_btn",
    )


# =============================================================================
# Main Render Entry Point
# =============================================================================

def render():
    st.markdown(
        dash_title_bar(
            "Credit Summary",
            "PNONG Section — Overdue detail, availability merge, summary report per company",
        ),
        unsafe_allow_html=True,
    )

    if not st.session_state.get("data_processed", False):
        st.info("Please run the pipeline in 'Loading and Processing Data' first.")
        return

    df_overdue = st.session_state.get("df_overdue")
    if df_overdue is None or (hasattr(df_overdue, "empty") and df_overdue.empty):
        st.warning("df_overdue is empty. Please re-process the data pipeline.")
        return

    df_avail = st.session_state.get("df_avail")

    # ------------------------------------------------------------------
    # Section 1: Raw Overdue + Dynamic Filter
    # ------------------------------------------------------------------
    df_s1_filtered = _render_section1_raw_table(df_overdue)

    # ------------------------------------------------------------------
    # Section 2: Merged (Overdue + Availability) — As-of Join
    # ------------------------------------------------------------------
    _render_section2_merged_table(df_s1_filtered, df_avail)

    # ------------------------------------------------------------------
    # Section 3: Summary Report per Company + Statistics + Export
    # ------------------------------------------------------------------
    _render_section3_summary(df_s1_filtered)